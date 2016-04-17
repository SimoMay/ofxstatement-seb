# -*- coding: utf-8 -*-

import re
import itertools
import logging

from datetime import datetime
from openpyxl import load_workbook

from ofxstatement.parser import StatementParser
from ofxstatement.plugin import Plugin
from ofxstatement.statement import Statement, StatementLine


def take(n, iterable):
    """Return first n items of the iterable as a list."""
    return list(itertools.islice(iterable, n))


def validate_workbook(workbook):
    """
    Naive validation to make sure that xlsx document is structured the way it was
    when this parser was written.

    :raises ValueError if workbook has invalid format
    """

    sheet = workbook.active
    try:
        rows = take(5, sheet.iter_rows())
        assert len(rows) == 5, 'Sheet should have at least 5 rows.'

        rows = [[c.value for c in row] for row in rows]

        header = rows[0]
        assert len(header[0])
        assert ['Saldo', 'Disponibelt belopp', 'Beviljad kredit', None, None] == header[1:]

        accounts = 1
        while not re.match(SebStatementParser.header_regexp, rows[accounts][0]):
            account_id = rows[accounts][0]
            logging.info('Detected account: %s' % account_id)
            accounts += 1
        logging.info('Total (%s) accounts detected.' % accounts)

        offset = 0 + accounts + 1

        header = rows[offset]
        assert re.match(SebStatementParser.header_regexp, header[0])
        assert [None, None, None, None, None] == header[1:]

        header = rows[offset + 1]
        assert re.match('^Bokförings\- *datum$', header[0])
        assert re.match('^Valuta\- *datum$', header[1])
        assert re.match('^Verifikations\- *nummer$', header[2])
        assert [None, None, None] == header[3:]

        header = rows[offset + 2]
        assert [None, None, None, 'Text / mottagare', 'Belopp', 'Saldo'] == header

    except AssertionError as e:
        raise ValueError(e)


class SebStatementParser(StatementParser):
    date_format = '%Y-%m-%d'
    bank_id = 'SEB'
    currency_id = 'SEK'
    header_regexp = '^Datum: ([0-9]{4}-[0-9]{2}-[0-9]{2}) - ([0-9]{4}-[0-9]{2}-[0-9]{2})$'

    @staticmethod
    def create(fin):
        wb = load_workbook(filename=fin, read_only=True)
        validate_workbook(wb)
        return SebStatementParser(wb)

    def __init__(self, workbook):
        self.workbook = workbook
        self.statement = self.parse_statement(workbook)

    def parse_statement(self, workbook):
        """
        Parse information from xlsx header that could be used to populate statement.

        :return: statment object
        """

        statement = Statement()
        sheet = workbook.active

        # We need only first 2 rows here.
        rows = take(3, sheet.iter_rows())
        rows = [[c.value for c in row] for row in rows]

        values = rows[1]
        privatkonto, saldo, disponibelt_belopp, beviljad_kredit, _1, _2 = values
        statement.account_id = privatkonto
        statement.end_balance = float(saldo)
        statement.bank_id = self.bank_id
        statement.currency = self.currency_id

        header = rows[2]
        m = re.match(self.header_regexp, header[0])
        if m:
            part_from, part_to = m.groups()
            statement.start_date = self.parse_datetime(part_from)
            statement.end_date = self.parse_datetime(part_to)

        return statement

    def split_records(self):
        sheet = self.workbook.active

        # Skip first 5 rows. Headers they are.
        for row in itertools.islice(sheet.iter_rows(), 5, None):
            # Row is potentially big so we yield generator.
            yield (c.value for c in row)

    def parse_record(self, row):
        row = take(5, row)

        stmt_line = StatementLine()
        stmt_line.date = self.parse_datetime(row[0])
        _ = self.parse_datetime(row[1])
        stmt_line.id = row[2]
        stmt_line.refnum = row[2]
        stmt_line.memo = row[3]
        stmt_line.amount = row[4]

        #
        # Looks like SEB formats description for card transactions so it includes the actual purchase date
        # within e.g. 'WIRSTRÖMS PU/14-12-31' and it means that description is 'WIRSTRÖMS PU' while the actual
        # card operation is 2014-12-31.
        #
        # P.S. Wirströms Irish Pub is our favorite pub in Stockholm.
        #
        m = re.match('(.*)/([0-9]{2}-[0-9]{2}-[0-9]{2})$', stmt_line.memo)
        if m:
            stmt_line.memo, date_string = m.groups()
            stmt_line.date_user = datetime.strptime(date_string, '%y-%m-%d')

        return stmt_line


class SebPlugin(Plugin):
    def get_parser(self, fin):
        return SebStatementParser.create(fin)
